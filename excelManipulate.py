import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def processWorkbook(filename):
	wb = xl.load_workbook(filename)
	sheet = wb['Sheet1']

	for row in range(2, sheet.max_row+1):
		price = sheet.cell(row,3)
		newPrice = price.value * 0.8
		newPriceCell = sheet.cell(row,4)
		newPriceCell.value = newPrice

	values = Reference(sheet,
				min_row=2,
				max_row=sheet.max_row,
				min_col=4,
				max_col=4)

	chart = BarChart()
	chart.add_data(values)
	sheet.add_chart(chart)

	wb.save(filename)
